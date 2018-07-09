import sys;

sys.path.append('.');
sys.path.append('dist-packages/');
import openpyxl, sys,subprocess,glob,base64;
import datetime, time;
import os,signal,getpass;
import logging,requests;

from cryptography.fernet import Fernet




s = signal.signal(signal.SIGINT, signal.SIG_IGN)
urlIP= 'http://10.20.3.193:8090';
filePath="DraftEnvStats.xlsx";
with open("ConflConfig.json",mode="r") as userFile:
		data=userFile.read().splitlines();

user=data[0];
password=data[1];
usermain=data[2];
userpass=base64.b64encode(user+":"+password);

myLogger = logging.getLogger('myapp')
handler = logging.FileHandler('envStatusLog.log')
formatter = logging.Formatter('%(asctime)s %(levelname)s %(message)s')
handler.setFormatter(formatter)
myLogger.addHandler(handler) 
myLogger.setLevel(logging.INFO)
myLogger.info('***********************************');
myLogger.info('***Starting Script***');

## Updating Status
def exithere(sheet):
	openmach=0;occupiedmach=0;errormach=0;
	row_count=sheet.max_row;
	coloumn_count=sheet.max_column;

	for icount in range(1,row_count):
		statusval=sheet.cell(row=icount+1,column=4).value;

		if statusval in ["Open","OPEN","open"]:
			openmach+=1;
		elif statusval in ["Occupied","OCCUPIED","occupied"]:
			occupiedmach+=1;
		elif statusval in ["Down","DOWN","down"]:
			errormach+=1;

	## Updating Status
	workbook=openpyxl.load_workbook(filePath)
	## Selecting the Sheet to edit
	sheet=workbook.get_sheet_by_name('Sheet1')
	sheet.cell(row=3,column=10).value=row_count-1
	sheet.cell(row=3,column=11).value=occupiedmach
	sheet.cell(row=3,column=12).value=openmach
	sheet.cell(row=3,column=13).value=errormach
	myLogger.info('Updating Status in Excel');
	workbook.save(filePath)
	#Updating to Excel Confluence 
	headers = {'X-Atlassian-Token': 'nocheck',
    'Authorization':'Basic %s'% userpass,}
	files = {'file': (filePath, open(filePath, 'rb')),}
	data={"comment":"testing rest"}
	response = requests.post(urlIP+'/rest/api/content/917506/child/attachment/att917508/data', data=data,headers=headers, files=files)

	signal.signal(signal.SIGINT, s);
	myLogger.info('Exiting Script');
	myLogger.info('***********************************');
	myLogger.info('');
	os.remove(filePath);
	print;
	sys.exit();


def relFromRequest(member):
	count=0;
	response = requests.get(urlIP+'/download/attachments/917510/DraftEnvRequest.xlsx',
		auth=(user,password))
	filePath="DraftEnvRequest.xlsx";
	file=open(filePath,"w");
	file.write(response.content);
	file.close();
	#listEnvs();
	filePath="DraftEnvRequest.xlsx";
	workbook=openpyxl.load_workbook(filePath)
	sheet=workbook.get_sheet_by_name('Sheet1')
	Row_count=sheet.max_row; 
	
	for count in range(Row_count):
		count=count+1;
		nameFExcel=(sheet.cell(row=count,column=3).value);#.strip();
		
		if (nameFExcel==member):
			print "**Removing Request from Request Sheet**";
			remRowAt=count; 
			sheet.cell(row=remRowAt,column=2).value="";
			sheet.cell(row=remRowAt,column=3).value="";
			sheet.cell(row=remRowAt,column=4).value="";
							
			workbook.save(filePath);
			headers = {
	  			'X-Atlassian-Token': 'nocheck',
			}
			files = {
		 		'file': (filePath, open(filePath, 'rb')),
			}
			data={"comment":"Release"}
			response = requests.post(urlIP+'/rest/api/content/917510/child/attachment/att917514/data', data=data,headers=headers, files=files, auth=(user, password))
			print "******Request Updated Sucessfully******";
			file.close();
			relexit=1;

	os.remove(filePath);
#End of release from Request Sheet




def main():
		#All Declarations
	path='.';	
## reading Config File
	with open("ConflConfig.json",mode="r") as userFile:
		data=userFile.read().splitlines();

	user=data[0];
	password=data[1];
	userpass=base64.b64encode(user+":"+password);

	headers={'Authorization':'Basic %s'% userpass}
	response = requests.get(urlIP+'/download/attachments/917506/DraftEnvStats.xlsx',headers=headers);
	filePath="DraftEnvStats.xlsx";
	file=open(filePath,"w");
	file.write(response.content);
	file.close();
	workbook=openpyxl.load_workbook(filePath)
	## Selecting the Sheet to edit
	sheet=workbook.get_sheet_by_name('Sheet1')
	myLogger.info('File imported successfully');
	servers=[];
	workingServers=[];
	foundTemp=0;
	INserver="";
	row_count = sheet.max_row;
	column_count = sheet.max_column;

	##Adding Servers to Array servers
	for row in range(1,row_count+1):
		servers.append(sheet.cell(row=row,column=2).value)
        print

	myLogger.info('Capturing Data');
	## For help parameter

	try:

		if (sys.argv[2]=="default" and sys.argv[4]=="default" and int(sys.argv[6])==0 and sys.argv[8]=="default" and sys.argv[9]=="--help"):
			print"********************************************************************************";
			print("Make arguments as:\n1. For Allocation of Environment: \n--name username(no spaces) --env environmentName --h hours(only hours in numeric)");
			print("(Eg. --name MAHESH --env CUT11 --h 4)");print
			print("2. Change Status of Environment: \n--change status(o:OPEN,d:DOWN,r:RELEASE) --env environmenName");
			print("(Eg. --change d --env CUT12)");
			print ("--name MAHESH --env CUT11 --h 4 --change d ");
			print"********************************************************************************";
			sys.exit();
		
		
	## For Allocation parameter**************************		
		#if (sys.argv[1]=="--name" and sys.argv[3]=="--env" and sys.argv[5]=="--h" ):
		if (sys.argv[2]!="default" and sys.argv[4]!="default" and int(sys.argv[6])!=0 and sys.argv[8]=="default" and sys.argv[9]!="--help"):
			changeOpSuccess=0;
			member=(sys.argv[2])
			if (member.isdigit())== True:
				print "ERROR!!! name format is incorrect!!!";
				myLogger.error("ERROR!! "+member+" name format is incorrect!!!");
				exithere(sheet);
			else:
				member=(sys.argv[2]).upper();
				duration=(sys.argv[6]); 
				if (duration.isdigit())==False:
					print "ERROR!!! hour format is incorrect!!!";
					myLogger.error("ERROR!! "+duration+" format is incorrect!!!");
					exithere(sheet);
				else:	
					duration=int(duration)
					###Updating Environment Status with Allocation (no parameters)
					myLogger.info('------In Member Environment Allocation');
					
					serverRequest=sys.argv[4]
					serverRequest=serverRequest.upper();
					for icount in range(1,len(servers)):
						statusVal=0;
						statusVal=str(sheet.cell(row=icount+1,column=4).value);
			    			if serverRequest==servers[icount] and statusVal in ["open","Open","OPEN"]:
	        					serverRequest=servers[icount];
	        					foundTemp=1;break;
					if foundTemp!=1:
			        		print 'Error!!! Environment is not in the List of OPEN Environment!!!';
		        			myLogger.error('Error!!! Environment is not in the List of OPEN Environment!!!');
	        				exithere(sheet);
			
					start_Time=time.strftime("%I:%M%p %d-%m-%Y",time.localtime(time.time()));
					end_Time=time.strftime("%I:%M%p %d-%m-%Y",time.localtime(time.time()+duration*3600));
			
					print "\n*****************************************"
					print "Member name         :",member;
					print "Allocate Environment:",serverRequest;
					print "Start Time          :",start_Time;
					print "End Time            :",end_Time;
					print "\n*****************************************"
			
					print "Updating Excel..."
					myLogger.info('Env Allocation %s--%s--%s'%(member,serverRequest,duration));
	        		##sheet.cell(row=icount+1,column=3).value=""
		        	sheet.cell(row=icount+1,column=4).value="OCCUPIED";
	        		sheet.cell(row=icount+1,column=5).value=member;
		        	sheet.cell(row=icount+1,column=6).value=start_Time;
	        		sheet.cell(row=icount+1,column=7).value=end_Time;
	        		workbook.save(filePath)
	        		print "*****Excel Updated*****";
		        	##Calling requestExcel Update
	        		relFromRequest(member);
	        		exithere(sheet);
				
		#elif (sys.argv[1]=="--change" and sys.argv[3]=="--env" and len(sys.argv)==5):		
		elif (sys.argv[2]=="default" and sys.argv[4]!="default" and int(sys.argv[6])==0 and sys.argv[8]!="default" and sys.argv[9]!="--help"):
			###FOR STATUS Change of Servers
			myLogger.info('In Environment Status Change');
			
			acceptStatus=sys.argv[8];
			INserver=sys.argv[4];INserver=INserver.upper();
			innCount=0;	changeOpSuccess=0;
			if acceptStatus not in ["O","o","D","d","R","r"]:
				print "ERROR!!!parameter for --change is incorrect.";
				exithere(sheet);
			

			print "********Changing Status of a Environment to DOWN/OPEN/RELEASE********\n";
			###FOR STATUS to DOWN to OPEN
			if acceptStatus in ["O","o"]:
				myLogger.info('------In DOWN->OPEN Servers');				
				for innCount in range(1,len(servers)):
					serStatus=sheet.cell(row=innCount+1,column=4).value;
					if INserver==sheet.cell(row=innCount+1,column=2).value and serStatus in ["DOWN","Down","down"]:
						sheet.cell(row=innCount+1,column=4).value="OPEN";
						print "DONE ...Status of %s Changed to OPEN..."%(INserver);
						myLogger.info('------%s DOWN->OPEN Changed'%(INserver));
						workbook.save(filePath);
						changeOpSuccess=1;
						exithere(sheet);

			###FOR STATUS to OPEN to DOWN
			elif acceptStatus in ["D","d"]:
				myLogger.info('------In OPEN->DOWN Servers');		
				for innCount in range(1,len(servers)):
					serStatus=sheet.cell(row=innCount+1,column=4).value;
					if INserver==sheet.cell(row=innCount+1,column=2).value and serStatus in ["OPEN","open","Open"]:
						sheet.cell(row=innCount+1,column=4).value="DOWN";
						print "DONE ...Status of %s Changed to DOWN..."%(INserver);
						myLogger.info('------%s OPEN->INACIVE Changed'%(INserver));
						workbook.save(filePath);
						changeOpSuccess=1;
						exithere(sheet); 

			###FOR STATUS to Occupied to OPEN / RELEASE
			elif acceptStatus in ["R","r"]:
				myLogger.info('------In OCCUPIED->OPEN Servers');
				for innCount in range(1,len(servers)):
					serStatus=sheet.cell(row=innCount+1,column=4).value;
					if INserver==sheet.cell(row=innCount+1,column=2).value and serStatus in ["OCCUPIED","Occupied","occupied"]:
						lastUsedBy=sheet.cell(row=innCount+1,column=5).value+" "+sheet.cell(row=innCount+1,column=6).value;
						sheet.cell(row=innCount+1,column=4).value="OPEN";
						sheet.cell(row=innCount+1,column=5).value="";
						sheet.cell(row=innCount+1,column=6).value="";
						sheet.cell(row=innCount+1,column=7).value="";
						sheet.cell(row=innCount+1,column=8).value=lastUsedBy;
						print "Status of %s Changed to OPEN..."%(INserver);
						myLogger.info('------%s OCCUPIED->OPEN Changed'%(INserver));
						workbook.save(filePath);
						changeOpSuccess=1;
						exithere(sheet);

			### exit else statement for change Operation
			if (changeOpSuccess==0):
				print "ERROR!!!"+INserver+" is not a Environment from available List!!!";
				myLogger.error('Error!!! Environment is not Found in the List !!!');
				exithere(sheet);


			## For Environment Status Change parameter**************************
		else:
			print "ERROR!!!parameter found incorrect, Please check --help.";
			exithere(sheet);

	except IndexError:
		print "Error!!! Wrong/Incomplete set of Parameters..";
    	myLogger.error('Error!!! Wrong/Incomplete set of Parameters. Try --help ');
    	exithere(sheet);



main();
##***End of File***

			
