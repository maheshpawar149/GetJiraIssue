import sys;'''
This Script is Created for Environment Team to allocate , Release & Change status of Environments into CLS Confluence Page.
This Script read ,Edit & Upload Excel of Confluence Page using Rest API's of Confluence Application.
'''
import sys;
sys.path.append('modules');
sys.path.append('modules/dist-packages');
import openpyxl,sys,subprocess,glob,base64;
import datetime, time,json;
import os,signal,getpass;
import logging,requests;
from datetime import datetime;
from datetime import timedelta;

s = signal.signal(signal.SIGINT, signal.SIG_IGN)

urlIP= 'http://10.20.3.186:8091'; #URL for Confluence Server
nPath='modules/'; #New download path for Excel Sheets 
global Jira_Tickets;
Jira_Tickets=["CEA-1","CEA-6","CEA-7","CEA-8","CEA-9","CEA-10","CEA-11","CEA-12","CEA-13"];
global server;
server="http://10.20.3.0:8181";
timeZonecal="";


global fmt;
fmt='%H:%M %d-%m-%y';
#"%I:%M%p %d-%m-%Y"
filePath1=nPath+"CLSNow_Dev_Environment_Allocation_Status.xlsx"; #relative pathname for  Excel
## Starting logging Method & variable declaration

myLogger = logging.getLogger('myapp')
handler = logging.FileHandler('envStatusLog.log')#File path for Logs
formatter = logging.Formatter('%(asctime)s %(levelname)s %(message)s')
handler.setFormatter(formatter)
myLogger.addHandler(handler) 
myLogger.setLevel(logging.INFO)
myLogger.info('***********************************');
myLogger.info('***Starting Script***');

def exithere(sheet):
	#For Existing Script & Saving all data to Allocation Status Sheet
	#row_count=sheet.max_row;
	#coloumn_count=sheet.max_column;
	## Updating Status
	workbook1=openpyxl.load_workbook(filePath1)
	sheet=workbook1.get_sheet_by_name('Sheet1')	
	myLogger.info('Updating Status in Excel');
	workbook1.save(filePath1)
	#Updating to Excel Confluence 
	headers = {'X-Atlassian-Token': 'nocheck','Authorization':'Basic %s'% userpass,}
	files = {'file': (filePath1, open(filePath1, 'rb')),}
	logDate=time.strftime(fmt,time.localtime(time.time()));
	data=""
	response = requests.post(urlIP+'/rest/api/content/3178500/child/attachment/att3178508/data', data=data,headers=headers, files=files)

	signal.signal(signal.SIGINT, s);
	myLogger.info('Exiting Script');
	myLogger.info('***********************************');
	myLogger.info('');
	#os.remove(filePath1);
	print;
	print "Script is Existing Now!!!";
	#sys.exit();
##End of exithere()

def updatejira(Jira_Ticket,dataC,dataD,dataT,dataDC):
	headers={'Authorization':'Basic YWRtaW46Qm9uZUJvbmUxJA==','content-type' : 'application/json'} 
	print "in update";
	dataC=json.dumps(dataC);
	dataD=json.dumps(dataD);
	dataT=json.dumps(dataT);
	dataDC=json.dumps(dataDC);
	##for decription
	response = requests.put(server+'/rest/api/2/issue/'+Jira_Ticket,headers=headers,data=dataD);
	print response.content;	

	##for comments
	if (dataC!=""):
		print "Adding Comment"
		response = requests.post(server+'/rest/api/2/issue/'+Jira_Ticket+'/comment',headers=headers,data=dataC);
		#print response.content;
		
	##for transition
	if (dataT!=""):
		print "$$$$$Changing Status"
		response = requests.post(server+'/rest/api/2/issue/'+Jira_Ticket+'/transitions?expand=transitions.fields',headers=headers,data=dataT);
		#print response.content;
	##FOR Update ENV after Allocation
	if (dataDC!=""):
		print "Changing Env "
		response = requests.put(server+'/rest/api/2/issue/'+Jira_Ticket,headers=headers,data=dataDC);
		#print response.content;
	
	#break;
	
def ConfReader():
	global durationLimit,userpass, multiple_extention,extention_on_req_present;
	from ConfigParser import SafeConfigParser
	from cryptography.fernet import Fernet	
	parser=SafeConfigParser()
	parser.read('Env_All.conf')
	myLogger.info('***Reading Config File***');
	##For durationLimit
	durationLimit=parser.get('Duration_Limit','durationLimit');	
	##For Authentication Segment
	cipher_key=parser.get('Login','key');
	encrypted_text=parser.get('Login','Segment');
	cipher = Fernet(cipher_key)
	decrypted_text = cipher.decrypt(encrypted_text)#reading cipher code
	userpass=base64.b64encode(decrypted_text);# base64 Conversion
	##For method Extension env duration
	multiple_extention=parser.get("Method_Overload","multiple_extention");
	extention_on_req_present=parser.get("Method_Overload","extention_on_req_present");
	
	myLogger.info('***Config File Reading Completed***');
#Calling Config Reader for Config Values	
ConfReader();

def callENVStatusEXCEl():
	#print str(sys.argv);
	##Get data from Environment Allocation Status Sheet
	headers={'Authorization':'Basic %s'% userpass}
	response = requests.get(urlIP+'/download/attachments/3178500/CLSNow_Dev_Environment_Allocation_Status.xlsx',headers=headers);
	filePath1=nPath+"CLSNow_Dev_Environment_Allocation_Status.xlsx";
	file=open(filePath1,"wb");
	file.write(response.content);	
	file.close();
	workbook1=openpyxl.load_workbook(filePath1);
	
	## Selecting the Sheet to edit
	sheet=workbook1.get_sheet_by_name('Sheet1')
	myLogger.info('File imported successfully');
	servers=[];	workingServers=[];
	foundTemp=0;
	INserver="";
	row_count = sheet.max_row;
	column_count = sheet.max_column;
	return workbook1;

def callENVReqEXCEL1():
	servers=[];foundTemp=0;
	##Getting Usage Sheet 
	headers={'Authorization':'Basic %s'% userpass}
	response = requests.get(urlIP+'/download/attachments/3178502/CLSNow_Dev_Environment_Usage.xlsx',headers=headers)
	filePath2=nPath+"CLSNow_Dev_Environment_Usage.xlsx";
	file=open(filePath2,"wb");
	file.write(response.content);
	file.close();
	workbook2=openpyxl.load_workbook(filePath1)
	sheet=workbook2.get_sheet_by_name('Sheet1')
	Row_count=sheet.max_row;
	myLogger.info('---Info!!! In Request Sheet Found!!');
	return workbook2;
	
def saveENVReqExcel1(workbook2):
			
	#Usage Sheet Updated with new values
	headers={'Authorization':'Basic %s'% userpass}
	response = requests.get(urlIP+'/download/attachments/3178502/CLSNow_Dev_Environment_Usage.xlsx',headers=headers);			
	workbook2.save(filePath2);
	headers = {'X-Atlassian-Token': 'nocheck','Authorization': 'Basic %s'% userpass,}
	files = {'file': (filePath2, open(filePath2, 'rb')),}
	data={"comment":"Env " +serverRequest+ " Allocated to "+memVal+"."}
	#data=[]
	response = requests.post(urlIP+'/rest/api/content/3178502/child/attachment/att3178503/data', data=data,headers=headers, files=files);
	print "******Request Updated Successfully******";
	#os.remove(filePath2);
	file.close();			
	
def catSelection(serverRequest):
	if (serverRequest in["CLSNOW","CCP","SANCTION","TEST"]):
		ReqPrio="CatAllocate";
		#print "--IN CAT Section";
		if (serverRequest =="CLSNOW"):
			#print "---IN CAT CLSNOW"
			serverRequest=["CUT12","CUT13","CUT14"];							
		elif (serverRequest =="CCP"):
			#print "---IN CAT CCP"
			serverRequest=["CUT05"];
		elif (serverRequest =="SANCTION"):
			#print "---IN CAT SANCTION"
			serverRequest=["CUT02"];
		elif (serverRequest =="TEST"):
			#print "---IN CAT TEST"
			serverRequest=["TEST1","TEST2","TEST3"];
	
	print serverRequest;
	return 	serverRequest;
					
## Updating Status

def readRequests(reqFrom,envVal):
	#callENVReqEXCEL();
	print;
	print "*****In Read REQUEST";
	servers=[];foundTemp=0;
	##Getting Usage Sheet 
	headers={'Authorization':'Basic %s'% userpass}
	response = requests.get(urlIP+'/download/attachments/3178502/CLSNow_Dev_Environment_Usage.xlsx',headers=headers)
	filePath2=nPath+"CLSNow_Dev_Environment_Usage.xlsx";
	file=open(filePath2,"wb");
	file.write(response.content);
	file.close();
	workbook2=openpyxl.load_workbook(filePath2)
	sheet=workbook2.get_sheet_by_name('Sheet1')
	Row_count=sheet.max_row;
	myLogger.info('---Info!!! In Request Sheet Found!!');
	
	for rc in range(1,Row_count+1):
		#if((sheet.cell(row=row,column=10).value)!='COMPLETED'):
		servers.append(sheet.cell(row=rc,column=2).value)
	#print servers
	exServerRequest="";
	
	for count in range(3,len(servers)+1):
		#RmemVal=(sheet.cell(row=count,column=4).value).upper();
		#RserverRequest=(sheet.cell(row=count,column=2).value).upper();
		statAll=sheet.cell(row=count,column=10).value;
		
		if( statAll=="REQUESTED" or statAll=="PENDING"):
			memVal=sheet.cell(row=count,column=4).value;
			teamVal=sheet.cell(row=count,column=3).value;
			serverRequest=sheet.cell(row=count,column=2).value;	
			durVal=sheet.cell(row=count,column=5).value;
			Jira_Ticket=sheet.cell(row=count,column=6).value;
			timezone=sheet.cell(row=count,column=9).value;
			print "Request Pass 1"
			print "To AllocateENV"
			print serverRequest; print "!!!";
			print sheet.cell(row=count,column=1).value;
			if (extention_on_req_present=="True"):
				print "extention_on_req_present TRUE"
				if (reqFrom=="fromExtend"):
					exServerRequest=catSelection(envVal);
					print exServerRequest;
					print "******************"
					if(serverRequest in exServerRequest):
						return "ReqFOUND";
				else:
					print "ALLOCATING NOW"
					allocateENV(Jira_Ticket,memVal,teamVal,serverRequest,durVal,timezone,servers,"fromRequest");			
					print "back From AllocateENV";print;
			else:
				print "ALLOCATING NOW"
				allocateENV(Jira_Ticket,memVal,teamVal,serverRequest,durVal,timezone,servers,"fromRequest");			
				print "Back From AllocateENV";print;
		

def allocateENV(Jira_Ticket,memVal,teamVal,envVal,durVal,timezone,servers,ReqPrio):
	print  "============IN ENV ALLOCATE";
	print envVal;
	
	foundTemp=0;
	workbook1=callENVStatusEXCEl();	
	sheet=workbook1.get_sheet_by_name('Sheet1')
	
	servers=[];
	Row_count=sheet.max_row;
	for rc in range(1,Row_count+1):
		#if((sheet.cell(row=row,column=10).value)!='COMPLETED'):
		servers.append(sheet.cell(row=rc,column=2).value)
	#print servers
		
	##If Parameter validation for member name & Team name
	if (memVal.isdigit())== True and (teamVal.isdigit())== True:
		print "ERROR!!! name format is incorrect!!!";
		myLogger.error("ERROR!! "+memVal+"or "+teamVal+" name format is incorrect!!!");
		exithere(sheet);
	else:
		if(int(durVal) >int(durationLimit)):
			error_msg= "Error!!!Duration should not be greater than %s hours"%(durationLimit);	
			print error_msg;
			myLogger.error("Error!!!Duration should not be greater than %s hours"%(durationLimit));	
			dataC={"body": "||{color:red}%s(x){color}||\n FROM: %s,%s,\nENV: %s\nFOR:%s hrs"%(error_msg,memVal,teamVal,serverRequest,durVal),}
			dataD={"update": {"description": [{"set": "||{color:red}%s(x){color}||\n\n |h5. FROM: | |h6. %s,%s,|\n|h5. ENV: | |h6. %s|\n|h5. FOR: | |h6. %s hrs|"%(error_msg,memVal,teamVal,serverRequest,durVal),}]}}
			dataT={"transition": {"id":"131"}}
			updatejira(Jira_Ticket,dataC,dataD,dataT);
			exithere(sheet);
			
		if not (timezone in ["IST","EST"]):
			print "Error!!!Timezone could be IST or EST";
			print "Error!!!Timezone could be IST or EST %s "%(timezone);
			exithere(sheet);

		if (durVal)==False:
			print "ERROR!!! hour format is incorrect!!!";
			myLogger.error("ERROR!! "+durVal+" format is incorrect!!!");
			exithere(sheet);
		else:
			durVal=int(durVal)
			myLogger.info('------In Member Environment Allocation');

			if (envVal=="DEFAULT"):
				##For AVAILABLE Env Selection
				for icount in range(1,len(servers)):
					statusVal=0;
					statusVal=str(sheet.cell(row=icount+1,column=4).value);
					if  statusVal in ["available","Available","AVAILABLE"]:
						statusVal=str(sheet.cell(row=icount+1,column=4).value);
						serverRequest=str(sheet.cell(row=icount+1,column=2).value);
						#print serverRequest;
						foundTemp=1;break;
			else:
				serverRequest=envVal
				serverRequest=serverRequest.upper();
						
						
			##Reading If request Env is AVAILABLE or not
			#print "$$";print serverRequest;
			##serverRequest=catSelection(serverRequest);
			print serverRequest;
			##Here catsel code
			
			if (serverRequest in["CLSNOW","CCP","SANCTION","TEST"]):
				ReqPrio="CatAllocate";
				#print "--IN CAT Section";
				if (serverRequest =="CLSNOW"):
					#print "---IN CAT CLSNOW"
					serverRequest=["CUT12","CUT13","CUT14"];							
				elif (serverRequest =="CCP"):
					#print "---IN CAT CCP"
					serverRequest=["CUT05"];
				elif (serverRequest =="SANCTION"):
					#print "---IN CAT SANCTION"
					serverRequest=["CUT02"];
				elif (serverRequest =="TEST"):
					#print "---IN CAT TEST"
					serverRequest=["TEST1","TEST2","TEST3"];
			
					#return 	serverRequest;
				for cCount in range(0,len(serverRequest)):
					#print cCount;
					for icount in range(1,len(servers)):					
						#print icount;
						if (foundTemp!=1):							
							#print serverRequest;
							statusVal=0;									
							statusVal=str(sheet.cell(row=icount+1,column=4).value);
							#print serverRequest[cCount];print servers[icount];
							if serverRequest[cCount]==servers[icount] and statusVal in ["available","Available","AVAILABLE"]:
								print "Available Environment FOUND !!!.Environment %s is Allocated"%(servers[icount]);
								#print serverRequest[cCount];print servers[icount];
								serverRequest=servers[icount];
								print serverRequest;
								rowloc=icount;
								print rowloc;
								foundTemp=1;break;
								
						else:
							print "In LAST BREAK ALLC";
							#icount=len(servers);
							break;
				print icount;			
			else:	
				print "In OriENV";
				#print len(servers);
				for icount in range(1,len(servers)):
					#print servers[icount];
					statusVal=0;
					statusVal=str(sheet.cell(row=icount+1,column=4).value);
					#print sheet.cell(row=icount+1,column=4).value
					#print servers[icount];print statusVal
					if serverRequest==servers[icount] and statusVal in ["available","Available","AVAILABLE"]:
						print "ENV Sep Found";
						#print servers[icount];print statusVal
						serverRequest=servers[icount];
						rowloc=icount;
						foundTemp=1;break;	
			
			##Calculating Start Time & End Time for Allocation					
			utcnow=datetime.utcnow();	
			if (timezone=="IST"):
				#IST START TIME
				istStart=utcnow+timedelta(hours=5,minutes=30);
				istStart=istStart.strftime(fmt);
				#IST END TIME
				istEnd=datetime.strptime(istStart,fmt)
				istEnd=istEnd+timedelta(hours=durVal);
				istEnd=istEnd.strftime(fmt);
				start_Time=istStart;
				end_Time=istEnd;						
			elif(timezone=="EST"):
				#EST START TIME
				estStart=utcnow+timedelta(hours=-4);
				estStart=estStart.strftime(fmt);				
				#EST END TIME
				estEnd=datetime.strptime(estStart,fmt)
				estEnd=estEnd+timedelta(hours=durVal);
				estEnd=estEnd.strftime(fmt);
				start_Time=estStart;
				end_Time=estEnd;
			
			if foundTemp!=1:
				##If Requested Env is not in AVAILABLE Status
				print 'Error!!! Environment %s is not AVAILABLE.your Request is Queue!!!'%(serverRequest);
				myLogger.error('Error!!! Environment %s is not AVAILABLE.your Request is Queue!!!'%(serverRequest));
				return 'Error!!! Environment %s is not AVAILABLE.your Request is Queue!!!'%(serverRequest);
				#if (ReqPrio=="fromRequest"):
					#print "+++CallUpdateRequestFFRR";
					##updateRequestPage(memVal,teamVal,serverRequest,durVal," "," ",timezone,"REQUESTED",ReqPrio);
					#exithere(sheet);
				#else:
					#print "+++CallUpdateRequestNOTFFRR";
					#updateRequestPage(memVal,teamVal,serverRequest,durVal,start_Time,end_Time,timezone,"REQUESTED",ReqPrio);
					#exithere(sheet);
			else:
				print rowloc;
				print "\n*****************************************"
				print "Member name         :",memVal;
				print "Team name           :",teamVal;
				print "Allocate Environment:",serverRequest;
				print "Start Time          :",start_Time;
				print "End Time            :",end_Time;
				print "TimeZone            :",timezone;
				print "\n*****************************************"
				print "Updating Excel..."
				myLogger.info('Env Allocation %s--%s--%s'%(memVal,serverRequest,durVal));
				sheet.cell(row=rowloc+1,column=4).value="ALLOCATED";
				sheet.cell(row=rowloc+1,column=5).value=teamVal;
				sheet.cell(row=rowloc+1,column=6).value=memVal;
				sheet.cell(row=rowloc+1,column=7).value=start_Time;
				sheet.cell(row=rowloc+1,column=8).value=end_Time;
				sheet.cell(row=rowloc+1,column=9).value=timezone;
				workbook1.save(filePath1);
				print "*****Excel Updated*****";
				#Calling UpdateRequestpage for updating Request data in Request Page
				updateRequestPage(Jira_Ticket,memVal,teamVal,serverRequest,durVal,start_Time,end_Time,timezone,"ALLOCATED",ReqPrio);

				#Call JIRA UPDATE

				
				exithere(sheet);	
			
def updateRequestPage(Jira_Ticket,memVal,teamVal,serverRequest,durVal,start_Time,end_Time,timezone,ReqType,ReqPrio):
	#print ReqType;
	print "IN UPDATE REQUEST"
	myLogger.info('---Info!!! In Env Request Page Update');
	myLogger.info('---[DATA]!!! In Env Request Page Update %s %s %s %s %s %s %s %s'%(memVal,teamVal, serverRequest, durVal, start_Time, end_Time, timezone, ReqType));
	print('---[DATA]!!! In Env Request Page Update %s %s %s %s %s %s %s %s'%(memVal,teamVal, serverRequest, durVal, start_Time, end_Time, timezone, ReqType));
	#print memVal;print teamVal;print serverRequest;print durVal;print start_Time;print end_Time;print timezone;print ReqType;
	
	
	#workbook2=callENVReqEXCEL();
	servers=[];foundTemp=0;
	##Getting Usage Sheet 
	headers={'Authorization':'Basic %s'% userpass}
	response = requests.get(urlIP+'/download/attachments/3178502/CLSNow_Dev_Environment_Usage.xlsx',headers=headers)
	filePath2=nPath+"CLSNow_Dev_Environment_Usage.xlsx";
	file=open(filePath2,"wb");
	file.write(response.content);
	file.close();
	
	workbook2=openpyxl.load_workbook(filePath2)
	sheet=workbook2.get_sheet_by_name('Sheet1')
	Row_count=sheet.max_row;
	myLogger.info('---Info!!! In Request Sheet Found!!');		
	
	servers=[];
	serverRequestCat=serverRequest;
	if(ReqPrio=="CatAllocate"):		
		if (serverRequest in["CUT12","CUT13","CUT14"]):
			serverRequest="CLSNOW";
		elif (serverRequest in["CUT05"]):
			serverRequest="CCP";
		elif (serverRequest in["CUT02"]):
			serverRequest="SANCTION";
		elif (serverRequest in["TEST1","TEST2","TEST3"]):
			serverRequest="TEST";
		
	
	if ReqType=="REQUESTED":
		print "In ADD New"
		##For New Entry of ENV Allocation		
		count=Row_count+1;			
		sheet.cell(row=count,column=1).value=(count-2);
		sheet.cell(row=count,column=2).value=serverRequest.upper();		
		sheet.cell(row=count,column=3).value=teamVal.upper();
		sheet.cell(row=count,column=4).value=memVal.upper();
		sheet.cell(row=count,column=5).value=durVal;
		sheet.cell(row=count,column=6).value=Jira_Ticket;
		sheet.cell(row=count,column=7).value=start_Time;
		sheet.cell(row=count,column=8).value=end_Time;
		sheet.cell(row=count,column=9).value=timezone;
		sheet.cell(row=count,column=10).value=ReqType;
		#break;
	elif(ReqType=="COMPLETED" or ReqType=="REQUESTED"or ReqType=="ALLOCATED" or ReqType=="EXTENDED"):
		print "in Completed & extend"
		##For Update Request /RELEASE/EXTEND Entry of ENV Allocation
		for rc in range(1,Row_count+1):
			#if((sheet.cell(row=row,column=10).value)!='COMPLETED'):
			servers.append(sheet.cell(row=rc,column=2).value)
		#print servers
		for count in range(3,len(servers)+1):				
			RmemVal=(sheet.cell(row=count,column=4).value).upper();
			RserverRequest=(sheet.cell(row=count,column=2).value).upper();
			statAll=sheet.cell(row=count,column=10).value;	
			#print "##",count;
			#print (RmemVal,memVal);
			#print (RserverRequest,serverRequest,RmemVal,memVal);
			if( RmemVal==memVal.upper() and (RserverRequest==serverRequest.upper() and (statAll=="ALLOCATED"  or statAll=="REQUESTED" or statAll=="PENDING" or statAll=="EXTENDED"))):
				print "---in COMPLETED/REQUESTED/EXTENDED"
				
				print"****";print sheet.cell(row=count,column=1).value			
				sheet.cell(row=count,column=2).value=serverRequestCat;	
				#req on sheet.cell(row=count,column=6)=
				sheet.cell(row=count,column=7).value=start_Time;
				sheet.cell(row=count,column=8).value=end_Time;				
				if ReqType=="COMPLETED":
					sheet.cell(row=count,column=5).value=str(durVal)
					sheet.cell(row=count,column=10).value="COMPLETED";
					print "--in COMPLETED"
					break;
				elif (ReqType=="ALLOCATED" or ReqType=="REQUESTED"):
					print "--in ALLOCATED/COMPLETED"
					sheet.cell(row=count,column=5).value=str(durVal);
					sheet.cell(row=count,column=10).value="ALLOCATED";
				elif ReqType=="EXTENDED":
					sheet.cell(row=count,column=5).value=str(int(durVal)+int(sheet.cell(row=count,column=5).value));
					sheet.cell(row=count,column=10).value="EXTENDED";
					print "--in EXTENDED"
				break;
		#break;
	
	myLogger.info('[DATA] Request History %s||%s||%s||%s||%s||%s||%s'%(memVal,teamVal,serverRequest,durVal,start_Time,end_Time,ReqType));
	print ('[DATA] Request History %s||%s||%s||%s||%s||%s||%s'%(memVal,teamVal,serverRequest,durVal,start_Time,end_Time,ReqType));
	myLogger.info('Info!!! Request History Updated in Request Sheet');
	print "Request History Updated in Request Sheet.";	
	#saveENVReqExcel(workbook2);
	#Usage Sheet Updated with new values
	headers={'Authorization':'Basic %s'% userpass}
	response = requests.get(urlIP+'/download/attachments/3178502/CLSNow_Dev_Environment_Usage.xlsx',headers=headers);			
	workbook2.save(filePath2);
	headers = {'X-Atlassian-Token': 'nocheck','Authorization': 'Basic %s'% userpass,}
	files = {'file': (filePath2, open(filePath2, 'rb')),}
	data={"comment":"Env " +serverRequest+ " Allocated to "+memVal+"."}
	#data=[]
	response = requests.post(urlIP+'/rest/api/content/3178502/child/attachment/att3178503/data', data=data,headers=headers, files=files);
	print "******Request Updated Successfully******";
	

	#JIRA Update
	#if(ReqType=="COMPLETED"): #elif (ReqType=="REQUESTED"):
	dataC="";dataD="";dataT="";dataDC="";
	if (ReqType=="ALLOCATED"): 
		dataC={"body": "||{color:#008000}START USING ENV.(/){color}||\n\n FROM: %s,%s,\nENV: %s\nFOR:%s hrs"%(memVal,teamVal,serverRequest,durVal),}
		dataD={"update": {"description": [{"set": "||{color:#008000}REQUEST APPROVED START USING ENV.(/){color}||\n\n |h5. FROM: | |h6. %s,%s,|\n|h5. ENV: | |h6. %s|\n|h5. FOR: | |h6. %s hrs|"%(memVal,teamVal,serverRequest,durVal),}]}}
		dataT={"transition": {"id":"201"}}
		dataDC={"fields": {"customfield_10208": { "value": durVal }}};
		updatejira(Jira_Ticket,dataC,dataD,dataT,dataDC);
		
	elif (ReqType=="REQUESTED"): 
		dataC={"body": "||{color:#008000}REQUEST IS ADDED TO QUEUE.(/){color}||\n\n FROM: %s,%s,\nENV: %s\nFOR:%s hrs"%(memVal,teamVal,serverRequest,durVal),}
		dataD={"update": {"description": [{"set": "||{color:#008000}REQUEST IS ADDED TO QUEUE.(/){color}||\n\n |h5. FROM: | |h6. %s,%s,|\n|h5. ENV: | |h6. %s|\n|h5. FOR: | |h6. %s hrs|"%(memVal,teamVal,serverRequest,durVal),}]}}
		dataT={"transition": {"id":"201"}}
		updatejira(Jira_Ticket,dataC,dataD,dataT,dataDC);

	elif (ReqType=="EXTENDED"):
		print "Request Extension Request";
		dataC={"body": "||{color:#008000}EXTENSION APPROVED ,CONTINUE USING ENV. (/){color}||\n\n FROM: %s,%s,\nENV: %s\nFOR:%s hrs"%(memVal,teamVal,serverRequest,durVal),}
		dataD={"update": {"description": [{"set": "||{color:#008000}EXTENSION APPROVED ,START USING ENV. (/){color}||\n\n |h5. FROM: | |h6. %s,%s,|\n|h5. ENV: | |h6. %s|\n|h5. FOR: | |h6. %s hrs|"%(memVal,teamVal,serverRequest,durVal),}]}}
		dataT={"transition": {"id":"201"}}
		updatejira(Jira_Ticket,dataC,dataD,dataT,dataDC);

	elif (ReqType=="COMPLETED"):
		print "Env Released Request";
		dataC={"body": "||{color:#008000}ENV RELEASED. (/){color}||\n FROM: %s,%s,\nENV: %s\nFOR:%s hrs"%(memVal,teamVal,serverRequest,durVal),}
		dataD={"update": {"description": [{"set": "||{color:#008000}ENV RELEASED. (/){color}||\n\n |h5. FROM: | |h6. %s,%s,|\n|h5. ENV: | |h6. %s|\n|h5. USED: | |h6. %s hrs|"%(memVal,teamVal,serverRequest,durVal),}]}}
		dataT={"transition": {"id":"111"}}		
		updatejira(Jira_Ticket,dataC,dataD,dataT,dataDC);
		

	#os.remove(filePath2);
	file.close();			
##End of Environment Usage Sheet Update

#updateRequestPage('PRIYANKA_TAKAWALE','SETTELMENT_CONTROLLER','CUT14','3','15:34 30-04-18','18:34 30-04-18','IST','COMPLETED')
#sys.exit();
def helpDesc():
	print"\n********************************************************************************";
	#print ("""1. Allocating Env to Member:\nSelect Env & give value Environment Name as in (CUT01, CUT14).\nSelect Hour & give value in Whole Number as in (2, 5, 10).\nSelect MemberName & give value in Name as (Mahesh, Pravin, Asifali etc.).\nSelect TeamName & give value in Name as (CLSNOW,CLSNET etc.).""");
	#print("""\n2.Changing Environment Status:\nSelect Env & give value Environment Name as in (CUT01, CUT14).\nSelect change & give value change as in (o:AVAILABLE, m:MAINTENANCE, r:RELEASE).""");			
	#print("""\n3.Extend Environment Duration:\nSelect Env & give value Environment Name as in (CUT01, CUT14).\nEnter new Duration.\nSelect change & give value change as e.""");			
	print("Script Parameters:");
	print("1. For Allocate: 	Mahesh DEVOPS CUT11 3 default IST default");
	print("Ex. Env name=CUT11, cut12, CLSNOW::[CUT12,CUT13,CUT14],CCP::[CUT05],SANCTION::[CUT02]");print;
	print("2. For Release:  	default default TEST3 0 r default default");print;
	print("3. For Status Change:	default default CUT11 0 o default default");
	print("Ex. Status name=o,O[OPEN]/m,M[MAINTENANCE]");print;
	print("4. For Extending Duration: default default TEST3 4 e default default");print;	
	print"********************************************************************************";
	sys.exit();	
	

def main():
	#readRequests()
	workbook1=callENVStatusEXCEl();
	sheet=workbook1.get_sheet_by_name('Sheet1')
	
	global dataC,dataD,dataT,dataDC;
	dataC="";dataD="";dataT="";dataDC="";
	##JIRA
	headers={'Authorization':'Basic YWRtaW46Qm9uZUJvbmUxJA==','content-type' : 'application/json'} 
	#Jira_Tickets=['CEA-13']
	for i in range(len(Jira_Tickets)):				
		response = requests.get(server+'/rest/api/2/issue/'+Jira_Tickets[i]+'?expand=changelog',headers=headers);
		##response = requests.get('http://10.20.2.237:8181/rest/api/2/issue/Test-3?fields=description',headers=headers);
		json_data=json.loads(response.content);
		log_length=len(json_data['changelog']['histories']);
		#statusVal=json_data['changelog']['histories'][log_length-1]['items'][0]['toString'];
		memVal=str(json_data['changelog']['histories'][log_length-1]['author']['name']);
		Usermail=str(json_data['changelog']['histories'][log_length-1]['author']['emailAddress']);
		teamVal=str(json_data['fields']['customfield_10209']['name']);
		#RequestType=str(json_data['fields']['issuetype']['name']);
		durVal=json_data['fields']['customfield_10207'];
		serverRequest=str(json_data['fields']['customfield_10208']['value']);
		req_type=json_data['fields']['priority']['name'];

		##Arguments from Console/Bamboo
		
		memVal=(memVal.upper()).strip();
		teamVal=(teamVal.upper()).strip();
		envVal=(serverRequest.upper()).strip();
		durVal=int(durVal);
		#chagVal=(sys.argv[5].upper()).strip();
		chagVal="DEFAULT";


		# memVal=(memVal.upper()).strip();
		# teamVal=(teamVal.upper()).strip();
		# memVal="DEFAULT"
		# teamVal="DEFAULT"
		# envVal="CUT02"
		# durVal="DEFAULT"
		# chagVal="r"
		#chagVal="DEFAULT";


		#timezone=(sys.argv[6].upper()).strip();	
		timezone="IST";
		#helpVal=(sys.argv[7].upper()).strip();
		#print "printing argvs:"
		#print memVal; print teamVal;print envVal;print durVal;print chagVal;print timezone;print helpVal;
		#sys.exit()

		servers=[];
		row_count = sheet.max_row;	
		for row in range(1,row_count+1):
			servers.append(sheet.cell(row=row,column=2).value)
	        print

		myLogger.info('Capturing Data');

		print "************************";
		print Jira_Tickets[i];
		desc=str(json_data['fields']['description']);
		req=desc.split("|");
		req_type=req[3];
		print req_type;
		print "************************";


		# if (statusVal=="New Request for Environment"):
		# 	print "NEW Request";
		# 	data={"update": {"description": [{"set": "||{color:#008000}REQUEST IS IN PROCESS{color}||\n\n |h5. FROM: | |h6. %s,%s,|\n|h5. ENV: | |h6. %s|\n|h5. FOR: | |h6. %s hrs|"%(memVal,teamVal,serverRequest,durVal),}]}}			
		# 	updatejira(Jira_Tickets[i],dataC,"","");

		if (req_type=="ENV REQUEST"):
			print "Request Found";
			changeOpSuccess=0;
			#Calling ENVALLOCATE	
			print "Calling updateRequestPage";
			#allocateENV(memVal,teamVal,envVal,durVal,timezone,servers);
			updateRequestPage(Jira_Tickets[i],memVal,teamVal,envVal,durVal," "," ",timezone,"REQUESTED","fromMain");
			print "RETURN FROM updateRequestPage";
			#print "Calling ENVALLOCATE";
			#allocateENV(memVal,teamVal,envVal,durVal,timezone,servers,"fromMain";);
			#print "RETURN FROM ENVALLOCATE";
			print " TO ReadRequest";
			readRequests("fromNewAllc",envVal)
			print "RETURN FROM ReadRequest";
			exithere(sheet)	

		elif(req_type=="RELEASE REQUEST"):
			print "In Env Released Request";
			ReqCat="Release_Request";

			myLogger.info('In Environment Status Change');
			acceptStatus=chagVal;INserver=envVal;INserver=INserver.upper();
			innCount=0;	changeOpSuccess=0;
				
			###FOR STATUS to ALLOCATED to AVAILABLE / RELEASE
			myLogger.info('------In ALLOCATED->AVAILABLE Servers');
			for innCount in range(1,len(servers)):
				serStatus=sheet.cell(row=innCount+1,column=4).value;
				if INserver==sheet.cell(row=innCount+1,column=2).value and serStatus in ["ALLOCATED","Allocated","allocated"]:
					##Capturing data for request page update
					memVal=sheet.cell(row=innCount+1,column=6).value;
					teamVal=sheet.cell(row=innCount+1,column=5).value;
					serverRequest=sheet.cell(row=innCount+1,column=2).value;
					#durVal=0;
					newDurval=0;
					start_Time=sheet.cell(row=innCount+1,column=7).value;
					#end_Time=sheet.cell(row=innCount+1,column=8).value;

					timezone=sheet.cell(row=innCount+1,column=9).value;
					utcnow=datetime.utcnow();
					if (timezone=="IST"):
						#IST Completed TIME
						end_Time=utcnow+timedelta(hours=5,minutes=30);
					elif(timezone=="EST"):
						#EST Completed TIME
						end_Time=utcnow+timedelta(hours=-4);
					end_Time=end_Time.strftime(fmt);
					##Calculating Actual Duration of Environment Usage
					#start_Time="12:18 2018-04-20"
					#end_Time="15:18 2018-04-20"
					start_Time=datetime.strptime(start_Time,fmt);
					end_Time=datetime.strptime(end_Time,fmt);
					newDurval=end_Time-start_Time;	
					newDurval=str(newDurval);
					Dhours,Dmins,Ssecs=newDurval.split(":");
					newDurval="%s Hours & %s Mins"%(Dhours,Dmins);
					
					lastUsedBy=sheet.cell(row=innCount+1,column=5).value +"::"+sheet.cell(row=innCount+1,column=6).value+" "+str(end_Time);
					comDurL=lastUsedBy.split(" ");	
					newDurval= newDurval +"::Actual:"+comDurL[1]+" "+comDurL[2];

					sheet.cell(row=innCount+1,column=4).value="AVAILABLE";
					sheet.cell(row=innCount+1,column=5).value="";
					sheet.cell(row=innCount+1,column=6).value="";
					sheet.cell(row=innCount+1,column=7).value="";
					sheet.cell(row=innCount+1,column=8).value="";
					sheet.cell(row=innCount+1,column=9).value="";
					sheet.cell(row=innCount+1,column=10).value=lastUsedBy;
					print "Status of %s Changed to AVAILABLE..."%(INserver);
					myLogger.info('------%s ALLOCATED->AVAILABLE Changed'%(INserver));
					workbook1.save(filePath1);
					changeOpSuccess=1;

					updateRequestPage(Jira_Tickets[i],memVal,teamVal,serverRequest,newDurval,start_Time,end_Time,"","COMPLETED","fromRel");
						
					exithere(sheet);
					print "Calling ReadRequest"
					readRequests("fromReles",serverRequest);
					print "End ReadRequest"
					sys.exit();
			if(changeOpSuccess!=1):
				print "Warning!!!Environment %s not found as an ALLOCATED Environment."%(INserver);
				myLogger.warning("Warning!!!Environment %s not found as an ALLOCATED Environment."%(INserver));

		elif(req_type=="EXTEND REQUEST" or req_type=="FORCE EXTEND"):
			ReqCat="Extend_Request";
			print "Request Extension Request";
			acceptStatus=chagVal;INserver=envVal;INserver=INserver.upper();
			global totDurUsed;timezone="";
			changeOpSuccess=0;
			acceptStatus='e';
			if acceptStatus in ["E","e"]:
				if (req_type!="FORCE EXTEND"):
					isreqFound=readRequests("fromExtend",envVal);
				else:
					isreqFound="ReqFOUND not found";

				print isreqFound;print"****"
				if (isreqFound=="ReqFOUND"):
						print "Request for same Env is Found.No Extension is Authorized.";
						print "Env Released Request";
						dataC={"body": "||{color:red}NEED APPROVAL FOR EXTENSION (/){color}||\n FROM: %s,%s,\nENV: %s\nFOR:%s hrs"%(memVal,teamVal,serverRequest,durVal),}
						dataD={"update": {"description": [{"set": "||{color:red}NEED APPROVAL FOR EXTENSION (/){color}||\n\n |h5. FROM: | |h6. %s,%s,|\n|h5. ENV: | |h6. %s|\n|h5. USED: | |h6. %s hrs|"%(memVal,teamVal,serverRequest,durVal),}]}}
						dataT={"transition": {"id":"221"}}	

						updatejira(Jira_Tickets[i],dataC,dataD,dataT,dataDC)


				else:
					print 'here'			
					extendText="";extendFound=0;oldTODate=0;newToDate=0;
					myLogger.info('------In Extending the Duration of the environment');
					for innCount in range(1,len(servers)):
						serStatus= sheet.cell(row=innCount+2,column=4).value;
						if INserver==sheet.cell(row=innCount+2,column=2).value and serStatus in ["ALLOCATED","Allocated","allocated"]:
							print "IN EXTEND REQUEST STATUS"
							#print INserver; print sheet.cell(row=innCount+2,column=2).value
							oldTODate=sheet.cell(row=innCount+2,column=8).value;
							#Checking whether the Extension is 1st or not
							if(len(oldTODate)>=19):
								extendText,oldTODate=oldTODate.split("::");	
							#print oldTODate
							newToDate=datetime.strptime(oldTODate,fmt)
							newToDate=newToDate+timedelta(hours=int(durVal));
							newToDate=newToDate.strftime(fmt);
							print extendText;
						
							if(len(extendText)==0):
								print "IN SINGLE EXTENTION"
								newToDate="[ExtendedBy "+str(durVal)+" hours]::"+str(newToDate);
								totDurUsed=durVal;
								#break;
							else:
								print "IN MULTIPLE EXTENTION"
								print multiple_extention;
								if(multiple_extention=="False"):
									print "This Request has already used one Extension";
									print "Env Released Request";
									dataC={"body": "||{color:red}ONE EXTENSION IS ALREADY FULFILLED(/){color}||\n FROM: %s,%s,\nENV: %s\nFOR:%s hrs"%(memVal,teamVal,serverRequest,durVal),}
									dataD={"update": {"description": [{"set": "||{color:red}ONE EXTENSION IS ALREADY FULFILLED (/){color}||\n\n |h5. FROM: | |h6. %s,%s,|\n|h5. ENV: | |h6. %s|\n|h5. USED: | |h6. %s hrs|"%(memVal,teamVal,serverRequest,durVal),}]}}
									dataT={"transition": {"id":"201"}}	

									updatejira(Jira_Tickets[i],dataC,dataD,dataT,dataDC)
									exithere(sheet);
									sys.exit();
								elif(multiple_extention=="True"):
									exHours=extendText.split(" ");
									totDurUsed=(int(durVal)+int(exHours[1]));
									newToDate="[ExtendedBy "+str(totDurUsed)+" hours]::"+str(newToDate);
									#break;
						
							sheet.cell(row=innCount+2,column=8).value=newToDate;
							#print newToDate;
							print "Allocation Extended for %s for %s hours."%(INserver,durVal);
							myLogger.info("------Allocation Extended for %s for %s hours."%(INserver,durVal));
							
							workbook1.save(filePath1);
							memVal=sheet.cell(row=innCount+2,column=6).value;
							teamVal=sheet.cell(row=innCount+2,column=5).value;
							serverRequest=sheet.cell(row=innCount+2,column=2).value;
							#durVal="";
							start_Time=sheet.cell(row=innCount+2,column=7).value;
							#end_Time=sheet.cell(row=innCount+2,column=8).value;
							end_Time=newToDate;
							print end_Time;
							changeOpSuccess=1;
							updateRequestPage(Jira_Tickets[i],memVal,teamVal,serverRequest,durVal,start_Time,end_Time,timezone,"EXTENDED","fromExtend");
							exithere(sheet);			
					
					### exit else statement for change Operation
					if (changeOpSuccess==0):
						print "ERROR!!!"+INserver+" is not a Environment from Allocated ENV List!!!";
						myLogger.error("ERROR!!!"+INserver+" is not a Environment from Allocated ENV List!!!");
						exithere(sheet);


		else:
			print("No Update");
			#print "ERROR!!!parameter found incorrect, Please check --HELP.";
			#helpDesc();
			#exithere(sheet);


			
			
		##Adding Servers to Array servers
		
		
##		try:

			## For Allocation user with Environment **************************		
			#if (memVal!="DEFAULT" and teamVal!="DEFAULT" and envVal!="DEFAULT" and int(durVal)!=0  and chagVal=="DEFAULT" and helpVal=="DEFAULT" and timezone!="DEFAULT"):
		
main();

		##End of Try Block  for any Error/Exceptions
##		except IndexError:
##			print "Error!!! Wrong/Incomplete set of Parameters..";
##	    	myLogger.error('Error!!! Wrong/Incomplete set of Parameters. Try --HELP ');
##		helpDesc();	
	    	#exithere(sheet);
##End of main()

##***End of File***
